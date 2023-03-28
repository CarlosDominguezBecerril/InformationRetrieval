import json

import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import xlsxwriter 
import glob

from collections import defaultdict


def generate_excel(metrics, model_names, datasets):
    titles = ["model name", "similarity", "average"] + datasets
    wb = Workbook()

    # create sheets:
    for metric in metrics:
        ws1 = wb.create_sheet(title=metric)

        ws1["A1"] = metric
        for i, title in enumerate(titles):
            ws1[f"{xlsxwriter.utility.xl_col_to_name(i)}2"] = title
    
    del wb[wb.sheetnames[0]]

    row = 3
    for model_name in model_names:
        for metric in metrics:
            ws = wb[metric]
            print(f"{xlsxwriter.utility.xl_col_to_name(0)}{row}", model_name)
            ws[f"{xlsxwriter.utility.xl_col_to_name(0)}{row}"] = model_name
            ws[f"{xlsxwriter.utility.xl_col_to_name(0)}{row}"].alignment = Alignment(wrap_text=True)
            similarity = ""
            if "cos_sim" in model_name:
                similarity = "cos_sim"
            elif "dot_score" in model_name:
                similarity = "dot_score"
            ws[f"{xlsxwriter.utility.xl_col_to_name(1)}{row}"] = similarity

        path = f"output/{model_name}/results/"
        for dataset in datasets:
            results_path = f"{path}{dataset}_results.json"

            if not os.path.exists(results_path):
                print(f"Doesn't exist: {results_path}")
                continue

            with open(results_path) as json_file:
                results = json.load(json_file)
            
            # Get the index of the dataset
            index = titles.index(dataset)
            for metric in metrics:
                # Get the sheet
                ws = wb[metric]
                ws[f"{xlsxwriter.utility.xl_col_to_name(index)}{row}"] = results[metric]

        row += 1

    wb.save(filename="./results.xlsx")


if __name__ == "__main__":
    metrics = ["NDCG@10", "Recall@100", "MRR@10"]

    datasets = [
            "msmarco", "trec-covid", "nfcorpus", "nq", "hotpotqa", "fiqa", "arguana", 
            "webis-touche2020", "cqadupstack", "quora", "dbpedia-entity", "scidocs", 
            "fever", "climate-fever", "scifact"
    ]

    datasets.extend([f"cqadupstack_{dataset}" for dataset in ["android", "english", "gaming", "gis", "mathematica", "physics", "programmers", "stats", "tex", "unix", "webmasters", "wordpress"]])

    model_names = [
        "dot_score_msmarco_supervised_pretrained_on_distilbert-base-uncased",
        "dot_score_msmarco_cropping_pretrained_on_distilbert-base-uncased",
        "dot_score_msmarco_LLM_facebook_opt-125m_pretrained_on_distilbert-base-uncased",
        "dot_score_msmarco_LLM_facebook_opt-350m_pretrained_on_distilbert-base-uncased",
        "dot_score_msmarco_LLM_facebook_opt-1.3b_pretrained_on_distilbert-base-uncased",
        "dot_score_msmarco_LLM_facebook_opt-2.7b_pretrained_on_distilbert-base-uncased",
        "dot_score_msmarco_LLM_facebook_opt-6.7b_pretrained_on_distilbert-base-uncased",
        "dot_score_msmarco_LLM_facebook_opt-13b_pretrained_on_distilbert-base-uncased",
        "cos_sim_dot_score_msmarco_LLM_facebook_opt-13b_pretrained_on_distilbert-base-uncased",
        "",
        "cos_sim_msmarco_supervised_pretrained_on_distilbert-base-uncased",
        "cos_sim_msmarco_cropping_pretrained_on_distilbert-base-uncased",
        "cos_sim_msmarco_LLM_facebook_opt-125m_pretrained_on_distilbert-base-uncased",
        "cos_sim_msmarco_LLM_facebook_opt-350m_pretrained_on_distilbert-base-uncased",
        "cos_sim_msmarco_LLM_facebook_opt-1.3b_pretrained_on_distilbert-base-uncased",
        "cos_sim_msmarco_LLM_facebook_opt-2.7b_pretrained_on_distilbert-base-uncased",
        "cos_sim_msmarco_LLM_facebook_opt-6.7b_pretrained_on_distilbert-base-uncased",
        "cos_sim_msmarco_LLM_facebook_opt-13b_pretrained_on_distilbert-base-uncased",
        "",
        "cos_sim_msmarco_LLM_facebook_opt-13b_pretrained_on_bert-base-uncased",
        "",
        "cos_sim_msmarco_supervised_pretrained_on_distilbert-base-uncased",
        "domain_adaptation_LLM",
        "domain_adaptation_cropping",
        "",
    ]

    domain_adaptation = ["trec-covid", "nfcorpus", "nq", "hotpotqa", "fiqa", "arguana", "webis-touche2020", "quora", "dbpedia-entity", "scidocs", "fever", "climate-fever", "scifact"]
    domain_adaptation.extend([f"cqadupstack_{dataset}" for dataset in ["android", "english", "gaming", "gis", "mathematica", "physics", "programmers", "stats", "tex", "unix", "webmasters", "wordpress"]])

    for d in domain_adaptation:
        model_names.append(f"domain_adaptation/cos_sim_{d}_LLM_facebook_opt-2.7b_pretrained_on_cos_sim_ms_marco_supervised_pretrained_on_distilbert-base-uncased")
    
    for d in domain_adaptation:
        model_names.append(f"domain_adaptation/cos_sim_{d}_cropping_pretrained_on_cos_sim_ms_marco_supervised_pretrained_on_distilbert-base-uncased")

    model_names.extend([
        "dpr/cos_sim_dpr_msmarco_supervised_pretrained_on_distilbert-base-uncased",
        "dpr/dot_score_dpr_msmarco_supervised_pretrained_on_distilbert-base-uncased"
    ])

    generate_excel(metrics, model_names, datasets)