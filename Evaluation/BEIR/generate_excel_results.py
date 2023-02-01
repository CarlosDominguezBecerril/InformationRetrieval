import json

import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import glob

from collections import defaultdict

# Available metrics
# K = [1, 3, 5, 10, 100, 1000] // For all the metrics is available this K
# NDCG@K 
# MAP@K
# Recall@K
# P@K 
# MRR@K
# R_cap@K
# Hole@K

METRICS = ["NDCG@10", "Recall@100", "MRR@10"]
NORMAL_TITLES = [["model name", 30], ["algorithm", 10], ["trained on", 20], ["loss - dpr", 10], ["average", 10]]
DATASETS = [["msmarco", 10], ["trec-covid", 10], ["nfcorpus", 10], ["nq", 10], ["hotpotqa", 10], ["fiqa", 10], ["arguana", 10], ["webis-touche2020", 10], ["cqadupstack", 10], ["quora", 10], ["dbpedia-entity", 10], ["scidocs", 10], ["fever", 10], ["climate-fever", 10], ["scifact", 10]]
DATASETS = [dataset + [i] for i, dataset in enumerate(DATASETS)]

DEFAULT_WIDTH = 10

def create_excel(workbook_path, force=False):
    if not os.path.exists(workbook_path) or force:
        wb = Workbook()

        # create sheets:
        for title in METRICS:
            ws1 = wb.create_sheet(title=title)

            # Adjust widths
            for i, width in enumerate(NORMAL_TITLES + DATASETS):
                if len(width) >= 2:
                    width = width[1]
                else:
                    width = DEFAULT_WIDTH
                ws1.column_dimensions[chr(ord('a') + i)].width = width

            ws1["A1"] = title
            for i, title in enumerate(NORMAL_TITLES + DATASETS):
                title = title[0]
                ws1[f"{chr(ord('a') + i)}2"] = title

        del wb[wb.sheetnames[0]]

        wb.save(filename=workbook_path)

def copy_baseline(workbook_path, baseline_path, force=False):
    create_excel(workbook_path, force)

    wb = load_workbook(filename=workbook_path)
    last_row = 3
    with open(baseline_path, "r") as json_file:
        baselines_data = json.load(json_file)

        for baseline in baselines_data:
            if baseline["model name"] == "test":
                continue

            for metric in METRICS:
                if metric not in baseline["metrics"]:
                    continue

                # Select worksheet
                ws = wb[metric]
                for i, title in enumerate(NORMAL_TITLES):
                    title = title[0]
                    if title in baseline:
                        position = f"{chr(ord('a') + i)}{last_row}"
                        ws[position] = baseline[title]
                        ws[position].alignment = Alignment(wrap_text=True)

                for j, dataset in enumerate(DATASETS, start=i+1):
                    dataset = dataset[0]
                    if dataset in baseline["metrics"][metric]:
                        position = f"{chr(ord('a') + j)}{last_row}"
                        ws[position] = baseline["metrics"][metric][dataset]
                        ws[position].alignment = Alignment(wrap_text=True)

            last_row += 1

    wb.save(filename=workbook_path)

    return last_row

def copy_results(workbook_path, models_path, last_row, order=None):
    
    def copy(file_name, last_row):
        model_name = file_name.split("/")[-1]
        # print(model_name, last_row)
        for dataset in DATASETS:
            dataset_name = dataset[0]
            column_position = 5 + dataset[2]
            if os.path.exists(f"{file_name}/{dataset_name}/") and os.path.exists(f"{file_name}/{dataset_name}/results.json"):
                with open(f"{file_name}/{dataset_name}/results.json", "r") as json_file:
                    results = json.load(json_file)
                    for metric in METRICS:
                        ws = wb[metric]

                        # Copy metadata
                        ws[f"A{last_row}"] = model_name
                        ws[f"B{last_row}"] = algorithm

                        # Copy result
                        ws[f"{chr(ord('a') + column_position)}{last_row}"] = results[metric]

    wb = load_workbook(filename=workbook_path)

    root_dir = list(glob.glob(f"{models_path}*"))

    names_to_algorithm = {"dpr": "inbatch", "contriever": "moco"}
    for folder_name in root_dir:
        model = folder_name.split("/")[-1]
        algorithm = names_to_algorithm[model]
        systems = list(glob.glob(f"{folder_name}/*"))

        if len(systems) == 0:
            continue

        if order is not None and model in order:
            for model_name, o in order[model].items():
                file_name = f"{folder_name}/{model_name}"
                copy(file_name, last_row + o)
            
            last_row += len(order[model])

        # Add the new systems
        for system in systems:
            model_name = system.split("/")[-1]
            if model_name in order[model]:
                continue
            copy(system, last_row)
            last_row += 1

    wb.save(filename=workbook_path)

def copy_all(workbook_path, models_path, baseline_path, force=False, order=None):
    last_row = copy_baseline(workbook_path, baseline_path, force)
    copy_results(workbook_path, models_path, last_row, order)

if __name__ == "__main__":
    order = defaultdict(dict)
    order["dpr"] = [
            "dpr-marco-supervised",
            "dpr-contriever-questions",
            "dpr-contriever-questions-documents",
            "dpr-opt-125m-do_sample=True-p=0.9",
            "dpr-opt-125m-do_sample=True-p=0.9_x5",
            "dpr-opt-1.3b-do_sample=True-p=0.9",
            "dpr-opt-30b-do_sample=True-p=0.9",
            "dpr-opt-125m",
            "dpr-opt-350m",
            "dpr-opt-1.3b",
            "scifact_dpr_format_pretrained_on_dpr_supervised_ms_marco",
            "scifact_do_sample=True_p=0.9_prompt=standard_query_generation_format_facebook_opt-30b_pretrained_on_dpr_supervised_ms_marco",
            "scifact_contriever_format_pretrained_on_dpr_supervised_ms_marco",
            "nfcorpus_dpr_format_pretrained_on_dpr_supervised_ms_marco",
            "nfcorpus_do_sample=True_p=0.9_query_generation_format_facebook_opt-30b_pretrained_on_dpr_supervised_ms_marco",
            "nfcorpus_contriever_format_pretrained_on_dpr_supervised_ms_marco",
            "scifact_do_sample=True_p=0.9_prompt=instruct-gpt_query_generation_format_instruct-gpt_pretrained_on_dpr_supervised_ms_marco",
            "scifact_do_sample=True_p=0.9_prompt=keywords_query_generation_format_facebook_opt-30b_pretrained_on_dpr_supervised_ms_marco",
            "scifact_do_sample=True_p=0.9_prompt=title_query_generation_format_facebook_opt-30b_pretrained_on_dpr_supervised_ms_marco",
            "scifact_with_negatives_dpr_format_pretrained_on_dpr_supervised_ms_marco"
        ]

    for key, value in order.items():
        order[key] = {name: i for i, name in enumerate(value)}

    copy_all("./results.xlsx", "./output/", "./baseline.json", force=True, order=order)
    
    